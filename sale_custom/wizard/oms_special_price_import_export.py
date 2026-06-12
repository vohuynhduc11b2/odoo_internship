from odoo import models, fields, api
from odoo.exceptions import UserError
import base64, io
import openpyxl
import xlsxwriter
from datetime import datetime, timedelta

# ===== Helpers ép kiểu an toàn =====
def _to_float(v, default=0.0):
    if v is None or v == '':
        return float(default)
    try:
        if isinstance(v, str):
            v = v.replace(',', '').strip()
        return float(v)
    except Exception:
        return float(default)

def _to_date(v):
    if hasattr(v, 'date'):  # datetime/date từ openpyxl
        return v.date() if isinstance(v, datetime) else v
    # Excel serial
    try:
        n = float(v)
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=n)).date()
    except Exception:
        pass
    # Chuỗi ngày
    if isinstance(v, str) and v.strip():
        s = v.strip().replace('/', '-')
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y', '%Y/%m/%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
    return None


class OmsSpecialPriceImportExportWizard(models.TransientModel):
    _name = 'oms.special.price.import.export.wizard'
    _description = 'Import/Export Special Price Wizard'

    special_price_id = fields.Many2one('oms.special.price', string='Giá đặc biệt', required=True, readonly=True)
    file = fields.Binary('File', attachment=True)
    filename = fields.Char('Tên file')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        active_id = self.env.context.get('active_id')
        if active_id:
            res['special_price_id'] = active_id
        return res

    # ===== EXPORT dùng ir.attachment (không cần field mới/upgrade) =====
    def action_export_template(self):
        self.ensure_one()
        sp = self.special_price_id

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {'in_memory': True})
        sh = wb.add_worksheet("Special Price")

        headers = ['ProductCode', 'ProductName', 'SpecialPrice', 'ValidFrom', 'ValidTo', 'Note']
        for c, h in enumerate(headers):
            sh.write(0, c, h)

        r = 1
        for line in sp.line_ids:
            sh.write(r, 0, line.item_code or '')
            sh.write(r, 1, line.item_id.name or '')
            sh.write(r, 2, line.special_price or 0.0)
            sh.write(r, 3, line.valid_from.strftime('%Y-%m-%d') if line.valid_from else '')
            sh.write(r, 4, line.valid_to.strftime('%Y-%m-%d') if line.valid_to else '')
            sh.write(r, 5, line.note or '')
            r += 1

        wb.close()
        output.seek(0)
        data_b64 = base64.b64encode(output.read())
        fname = f"GiaDacBiet_{(sp.customer_id.name or 'Customer').replace(' ', '_')}.xlsx"

        att = self.env['ir.attachment'].create({
            'name': fname,
            'datas': data_b64,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {'type': 'ir.actions.act_url', 'url': f'/web/content/{att.id}?download=true', 'target': 'self'}

    # ===== IMPORT: bỏ qua dòng lỗi, cuối cùng chỉ thông báo tóm tắt =====
    def action_import_file(self):
        self.ensure_one()
        if not self.file:
            return {
                'type': 'ir.actions.client', 'tag': 'display_notification',
                'params': {'title': 'Import giá đặc biệt', 'message': 'Bạn phải chọn file để import!',
                           'type': 'warning', 'sticky': False}
            }

        sp = self.special_price_id
        try:
            wb = openpyxl.load_workbook(
                filename=io.BytesIO(base64.b64decode(self.file)),
                read_only=True, data_only=True
            )
            ws = wb.active
        except Exception as e:
            return {
                'type': 'ir.actions.client', 'tag': 'display_notification',
                'params': {'title': 'Import giá đặc biệt', 'message': f'Lỗi đọc file: {e}',
                           'type': 'danger', 'sticky': True}
            }

        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return {
                'type': 'ir.actions.client', 'tag': 'display_notification',
                'params': {'title': 'Import giá đặc biệt', 'message': 'File trống hoặc thiếu hàng tiêu đề.',
                           'type': 'warning', 'sticky': False}
            }

        # map header -> index (lowercase, bỏ khoảng trắng)
        norm = lambda s: str(s).strip().lower().replace(' ', '') if s else ''
        hidx = {norm(v): i for i, v in enumerate(header) if v}

        def col(*names):
            for n in names:
                i = hidx.get(norm(n))
                if i is not None:
                    return i
            return None

        # Alias header chấp nhận
        c_item  = col('productcode', 'itemcode', 'mãhàng')
        c_name  = col('productname', 'itemname')  # optional
        c_price = col('specialprice', 'price', 'giặcbiệt')
        c_from  = col('validfrom', 'fromdate', 'từngày')
        c_to    = col('validto', 'todate', 'đếnsngày', 'đếnngày')
        c_note  = col('note', 'ghichu')

        if c_item is None:
            return {
                'type': 'ir.actions.client', 'tag': 'display_notification',
                'params': {'title': 'Import giá đặc biệt', 'message': "Thiếu cột 'ProductCode'/'ItemCode'.",
                           'type': 'warning', 'sticky': False}
            }

        Product = self.env['product.product'].sudo()

        lines_vals = []
        missing_codes, invalid_rows, duplicate_rows = [], [], []
        seen = set()
        rno = 1

        for row in rows:
            rno += 1
            code = row[c_item] if c_item is not None else None
            if not code or str(code).strip() in ('', 'None'):
                continue  # bỏ dòng trống

            prod = Product.search([('default_code', '=', str(code).strip())], limit=1)
            if not prod:
                missing_codes.append(f"Dòng {rno}: {code}")
                continue

            sp_value = _to_float(row[c_price], 0.0) if c_price is not None else 0.0
            valid_from = _to_date(row[c_from]) if c_from is not None else None
            valid_to   = _to_date(row[c_to]) if c_to is not None else None
            note = (row[c_note] if c_note is not None else '') or ''

            if not valid_from:
                valid_from = sp.valid_from
            if not valid_to:
                valid_to = sp.valid_to

            # validate
            if valid_from and valid_to and valid_from > valid_to:
                invalid_rows.append(f"Dòng {rno}: ValidFrom > ValidTo (mã {code})")
                continue
            if sp_value < 0:
                invalid_rows.append(f"Dòng {rno}: Giá đặc biệt âm (mã {code})")
                continue

            key = (prod.id, valid_from, valid_to)
            if key in seen:
                duplicate_rows.append(f"Dòng {rno}: Trùng trong file (mã {code})")
                continue
            seen.add(key)

            lines_vals.append((0, 0, {
                'item_id': prod.id,
                'special_price': sp_value,
                'valid_from': valid_from,
                'valid_to': valid_to,
                'note': note,
            }))

        valid_count = len(lines_vals)
        skipped_missing = len(missing_codes)
        skipped_invalid = len(invalid_rows)
        skipped_dup = len(duplicate_rows)
        skipped_total = skipped_missing + skipped_invalid + skipped_dup

        # chỉ ghi nếu có dòng hợp lệ; nếu không thì giữ nguyên
        if valid_count:
            sp.line_ids.unlink()
            sp.write({'line_ids': lines_vals})

        # Thông báo tóm tắt (show tối đa 5 dòng/nhóm để gọn)
        def preview(lst, label):
            if not lst:
                return ''
            head = "\n".join(lst[:5])
            more = f"\n… (+{len(lst)-5} dòng nữa)" if len(lst) > 5 else ""
            return f"{label} ({len(lst)}):\n{head}{more}"

        sections = [preview(missing_codes, "Không tìm thấy SP"),
                    preview(invalid_rows, "Dữ liệu không hợp lệ"),
                    preview(duplicate_rows, "Trùng trong file")]
        detail = "\n\n".join([s for s in sections if s])

        msg = f"Import hoàn tất. Dòng hợp lệ: {valid_count}"
        if skipped_total:
            msg += f" · Bỏ qua: {skipped_total} (không thấy SP: {skipped_missing}, lỗi: {skipped_invalid}, trùng: {skipped_dup})"
            if detail:
                msg += "\n\n" + detail

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import giá đặc biệt',
                'message': msg,
                'type': 'warning' if skipped_total else 'success',
                'sticky': bool(skipped_total),
            }
        }
