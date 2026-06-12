from odoo import models, fields, api
import base64, io
import openpyxl, xlsxwriter

def _to_int(v, default=0):
    try:
        if v is None or v == "": return int(default)
        if isinstance(v, str): v = v.replace(',', '').strip()
        return int(round(float(v)))
    except Exception:
        return int(default)

class OmsPromoApplyProductIEWizard(models.TransientModel):
    _name = 'oms.promo.apply.product.ie.wizard'
    _description = 'Import/Export Products for Promotion'

    promotion_id = fields.Many2one('oms.promotion', required=True, readonly=True)
    file = fields.Binary('File (.xlsx)')
    filename = fields.Char('Tên file')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        active_id = self.env.context.get('active_id')
        if active_id:
            res['promotion_id'] = active_id
        return res

    # ---- EXPORT (1 sheet, đúng tab "Sản phẩm áp dụng") ----
    def action_export(self):
        self.ensure_one()
        promo = self.promotion_id
        out = io.BytesIO()
        wb = xlsxwriter.Workbook(out, {'in_memory': True})
        sh = wb.add_worksheet("ApplyProducts")
        headers = ['ProductCode','ProductName','QtyFrom','QtyTo']
        for c,h in enumerate(headers): sh.write(0,c,h)
        r = 1
        for l in promo.apply_product_line_ids:
            sh.write(r,0, l.product_tmpl_id.default_code or '')
            sh.write(r,1, l.product_tmpl_id.name or '')
            sh.write(r,2, l.qty_from or 1)
            sh.write(r,3, l.qty_to or 999999)
            r += 1
        wb.close(); out.seek(0)
        data_b64 = base64.b64encode(out.read())
        fname = f"KM_{(promo.code or 'promo').replace(' ','_')}_ApplyProducts.xlsx"
        att = self.env['ir.attachment'].create({
            'name': fname, 'datas': data_b64,
            'res_model': self._name, 'res_id': self.id,
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {'type':'ir.actions.act_url','url':f'/web/content/{att.id}?download=true','target':'self'}

    # ---- IMPORT (chỉ cập nhật tab "Sản phẩm áp dụng") ----
    def action_import(self):
        self.ensure_one()
        if not self.file:
            return self._notify('Bạn phải chọn file để import!', 'warning')

        try:
            wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.file)), read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return self._notify(f'Lỗi đọc file: {e}', 'danger', sticky=True)

        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return self._notify('File trống hoặc thiếu tiêu đề.', 'warning')

        # map header -> index (lowercase, bỏ khoảng trắng)
        norm = lambda s: str(s).strip().lower().replace(' ','') if s else ''
        hidx = {norm(v): i for i,v in enumerate(header) if v}

        def col(*names):
            for n in names:
                i = hidx.get(norm(n))
                if i is not None: return i
            return None

        c_code = col('productcode','itemcode','mãhàng')
        c_from = col('qtyfrom','minqty','minq')
        c_to   = col('qtyto','maxqty','maxq')
        if c_code is None:
            return self._notify("Thiếu cột 'ProductCode'/'ItemCode'.", 'warning')

        ProductTmpl = self.env['product.template'].sudo()
        ApplyLine = self.env['oms.promotion.apply.product.line'].sudo()
        promo = self.promotion_id

        created=updated=skipped=0
        seen=set()
        for rno,row in enumerate(rows, start=2):
            code = row[c_code] if c_code is not None else None
            if not code or str(code).strip() in ('','None'):
                continue
            code = str(code).strip()

            qty_from = _to_int(row[c_from] if c_from is not None else 1, 1)
            qty_to   = _to_int(row[c_to] if c_to is not None else 999999, 999999)

            key=(code, qty_from, qty_to)
            if key in seen:
                skipped += 1; continue
            seen.add(key)

            tmpl = ProductTmpl.search([('default_code','=',code)], limit=1)
            if not tmpl:
                skipped += 1; continue

            # upsert: nếu đã có dòng cùng product -> cập nhật qty_from/qty_to
            ex = ApplyLine.search([('promotion_id','=',promo.id), ('product_tmpl_id','=',tmpl.id)], limit=1)
            vals = {'promotion_id': promo.id, 'product_tmpl_id': tmpl.id,
                    'qty_from': qty_from, 'qty_to': qty_to}
            if ex:
                ex.write(vals); updated += 1
            else:
                ApplyLine.create(vals); created += 1

        msg = f'Import xong. Tạo mới: {created}, Cập nhật: {updated}, Bỏ qua: {skipped}.'
        return self._notify(msg, 'success' if (created or updated) else 'warning')

    def _notify(self, message, type='info', sticky=False, title='Import/Export SP KM'):
        return {
            'type':'ir.actions.client','tag':'display_notification',
            'params': {'title': title, 'message': message, 'type': type, 'sticky': sticky}
        }
