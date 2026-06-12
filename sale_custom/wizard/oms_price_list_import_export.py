from odoo import models, fields, api
from odoo.exceptions import UserError
import base64, io
import openpyxl
import xlsxwriter
from datetime import datetime, timedelta

# ===== Helpers ép kiểu an toàn =====
def _to_float(v, default=0.0):
    if v is None or v == '':
        return float(default)
    try:
        if isinstance(v, str):
            v = v.replace(',', '').strip()
        return float(v)
    except Exception:
        return float(default)

def _to_int(v, default=0):
    return int(round(_to_float(v, default)))

def _to_bool(v):
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return bool(int(v))
    if isinstance(v, str):
        s = v.strip().lower()
        return s in ('1', 'true', 'x', 'yes', 'y')
    return False

def _to_date(v):
    # openpyxl: date/datetime -> trả trực tiếp
    if hasattr(v, 'date'):
        return v.date() if isinstance(v, datetime) else v
    # Excel serial
    try:
        n = float(v)
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=n)).date()
    except Exception:
        pass
    # Chuỗi ngày
    if isinstance(v, str) and v.strip():
        s = v.strip().replace('/', '-')
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y', '%Y/%m/%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
    return None


class OmsPriceListImportExportWizard(models.TransientModel):
    _name = 'oms.price.list.import.export.wizard'
    _description = 'Import/Export Price List Wizard'

    pricelist_id = fields.Many2one('oms.price.list', string='Bảng giá', required=True, readonly=True)
    file = fields.Binary('File', attachment=True)
    filename = fields.Char('Tên file')

    # dùng cho export (không compute để tránh bị rỗng khi tải)
    export_file = fields.Binary('Export File')
    export_filename = fields.Char('Export Filename')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        active_id = self.env.context.get('active_id')
        if active_id:
            res['pricelist_id'] = active_id
        return res

    # ===== EXPORT (headers theo API ItemPriceAUT) =====
    def action_export_template(self):
        self.ensure_one()
        pricelist = self.pricelist_id

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {'in_memory': True})
        sh = wb.add_worksheet("Price List")

        headers = ['ItemCode', 'PriceList', 'Price', 'ListName']
        for c, h in enumerate(headers):
            sh.write(0, c, h)

        r = 1
        for line in pricelist.line_ids:
            sh.write(r, 0, line.item_code or '')
            sh.write(r, 1, line.api_price_list_id or '')
            sh.write(r, 2, line.price or 0.0)
            sh.write(r, 3, line.price_frame_name or '')
            r += 1

        wb.close()
        output.seek(0)
        self.export_file = base64.b64encode(output.read())
        self.export_filename = f"Banggia_{pricelist.name.replace(' ', '_')}.xlsx"

        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content?model={self._name}&id={self.id}&field=export_file"
                   f"&download=true&filename={self.export_filename or 'price_list.xlsx'}",
            'target': 'self',
        }

    # ===== IMPORT (map alias header + ép kiểu an toàn) =====
    def action_import_file(self):
        self.ensure_one()
        if not self.file:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Import bảng giá',
                    'message': 'Bạn phải chọn file để import!',
                    'type': 'warning',
                    'sticky': False,
                }
            }
    
        pricelist = self.pricelist_id
    
        # Đọc workbook
        try:
            wb = openpyxl.load_workbook(
                filename=io.BytesIO(base64.b64decode(self.file)),
                read_only=True, data_only=True
            )
            ws = wb.active
        except Exception as e:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Import bảng giá',
                    'message': f'Lỗi đọc file: {e}',
                    'type': 'danger',
                    'sticky': True,
                }
            }
    
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Import bảng giá',
                    'message': 'File trống hoặc thiếu hàng tiêu đề.',
                    'type': 'warning',
                    'sticky': False,
                }
            }
    
        # map header -> index (lowercase, bỏ khoảng trắng)
        norm = lambda s: str(s).strip().lower().replace(' ', '') if s else ''
        hidx = {norm(v): i for i, v in enumerate(header) if v}
    
        def col(*names):
            for n in names:
                i = hidx.get(norm(n))
                if i is not None:
                    return i
            return None
    
        c_item   = col('productcode','itemcode','mãhàng')
        c_api_pl = col('pricelist')
        c_list   = col('listname')
        c_from   = col('fromdate')
        c_to     = col('todate')
        c_min    = col('minqty','minq')
        c_max    = col('maxqty')
        c_ptype  = col('pricety','pricetype','bảnggiá')
        c_price  = col('price','giá')
        c_invoic = col('isinvoic','isinvoice')
        c_group  = col('groupc','groupcodesolar')
        c_level  = col('levelcc','levelcode')
    
        Product = self.env['product.product'].sudo()
        Frame = self.env['oms.pricelist.frame'].sudo().with_context(active_test=False)
        frame_domain = [('active', '=', True)]
        if pricelist.category_id:
            frame_domain.append(('category_id', '=', pricelist.category_id))
        frames = Frame.search(frame_domain)
        ordered_frames = frames.sorted(lambda f: (f.min_qty or 0, f.max_qty or 0, f.price_list_name or '', f.id))
        bg_map = {frame.id: f"BG{idx:02d}" for idx, frame in enumerate(ordered_frames, start=1)}
    
        # Counters & logs
        lines_vals = []
        missing_codes, invalid_rows, duplicate_rows = [], [], []
        seen = set()
    
        rno = 1
        for row in rows:
            rno += 1
            code = row[c_item] if c_item is not None else None
            if not code or str(code).strip() in ('', 'None'):
                continue  # dòng trống
            
            prod = Product.search([('default_code', '=', str(code).strip())], limit=1)
            if not prod:
                missing_codes.append(f"Dòng {rno}: {code}")
                continue
            
            from_date = _to_date(row[c_from]) if c_from is not None else None
            to_date   = _to_date(row[c_to]) if c_to is not None else None
            if not from_date:
                from_date = pricelist.from_date
            if not to_date:
                to_date = pricelist.to_date
    
            frame = False
            if c_api_pl is not None and row[c_api_pl] not in (None, ''):
                frame = frames.filtered(lambda f: f.api_id == _to_int(row[c_api_pl], 0))[:1]
            if not frame and c_list is not None and row[c_list]:
                list_name = str(row[c_list]).strip()
                frame = frames.filtered(lambda f: (f.price_list_name or '').strip() == list_name)[:1]

            min_qty  = _to_int(row[c_min], 1) if c_min is not None else int(frame.min_qty or 1) if frame else 1
            max_qty  = _to_int(row[c_max], 9_999_999) if c_max is not None else int(frame.max_qty or 9_999_999) if frame else 9_999_999
            price_ty = (row[c_ptype] if c_ptype is not None else bg_map.get(frame.id, 'BG01') if frame else 'BG01') or 'BG01'
            price    = _to_float(row[c_price], 0.0) if c_price is not None else 0.0
            is_inv   = _to_bool(row[c_invoic]) if c_invoic is not None else False
            groupc   = (row[c_group] if c_group is not None else '') or ''
            levelc   = (row[c_level] if c_level is not None else '') or ''
    
            # validate
            if from_date and to_date and from_date > to_date:
                invalid_rows.append(f"Dòng {rno}: FromDate > ToDate (mã {code})")
                continue
            if min_qty > max_qty:
                invalid_rows.append(f"Dòng {rno}: MinQty > MaxQty (mã {code})")
                continue
            if price < 0:
                invalid_rows.append(f"Dòng {rno}: Giá âm (mã {code})")
                continue
            
            key = (prod.id, from_date, to_date, min_qty, max_qty, str(price_ty))
            if key in seen:
                duplicate_rows.append(f"Dòng {rno}: Trùng trong file (mã {code})")
                continue
            seen.add(key)
    
            lines_vals.append((0, 0, {
                'item_id': prod.id,
                'from_date': from_date,
                'to_date': to_date,
                'min_qty': min_qty,
                'max_qty': max_qty,
                'price_type': price_ty,
                'price': price,
                'price_frame_id': frame.id if frame else False,
                'price_frame_name': frame.price_list_name if frame else (row[c_list] if c_list is not None else ''),
                'is_invoice': is_inv,
                'group_code_solar': groupc,
                'level_code': levelc,
            }))
    
        valid_count = len(lines_vals)
        skipped_missing = len(missing_codes)
        skipped_invalid = len(invalid_rows)
        skipped_dup = len(duplicate_rows)
        skipped_total = skipped_missing + skipped_invalid + skipped_dup
    
        # Chỉ thay thế khi có dòng hợp lệ; nếu không, giữ nguyên dữ liệu cũ
        if valid_count:
            pricelist.line_ids.unlink()
            pricelist.write({'line_ids': lines_vals})
    
        # Ghép thông điệp ngắn gọn + một vài chi tiết
        details = []
        for lst, label in (
            (missing_codes, "Không tìm thấy SP"),
            (invalid_rows, "Dữ liệu không hợp lệ"),
            (duplicate_rows, "Trùng trong file"),
        ):
            if lst:
                preview = "\n".join(lst[:5])  # show tối đa 5 dòng để gọn
                more = f"\n… (+{len(lst)-5} dòng nữa)" if len(lst) > 5 else ""
                details.append(f"{label} ({len(lst)}):\n{preview}{more}")
    
        msg_header = f"Import hoàn tất. Dòng hợp lệ: {valid_count}"
        if skipped_total:
            msg_header += f" · Bỏ qua: {skipped_total} " \
                          f"(không thấy SP: {skipped_missing}, lỗi: {skipped_invalid}, trùng: {skipped_dup})"
    
        msg = msg_header
        if details:
            msg += "\n\n" + "\n\n".join(details)
    
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import bảng giá',
                'message': msg,
                'type': 'warning' if skipped_total else 'success',
                'sticky': bool(skipped_total),
            }
        }
