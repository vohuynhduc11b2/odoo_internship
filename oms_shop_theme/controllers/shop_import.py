import base64
import io
import logging

import openpyxl
import xlsxwriter

from odoo import http
from odoo.http import request

_logger = logging.getLogger(__name__)


class OmsShopImportController(http.Controller):

    def _empty_import_result(self):
        return {
            'success_count': 0,
            'failed_count': 0,
            'added_qty': 0.0,
            'cart_quantity': 0,
            'failures': [],
        }

    def _finish_import_result(self, order, result):
        try:
            order.website_apply_auto_purchase_promo()
        except Exception:
            _logger.exception('Apply promo failed after cart import')
        try:
            if hasattr(order, '_compute_amounts'):
                order._compute_amounts()
            elif hasattr(order, '_amount_all'):
                order._amount_all()
        except Exception:
            pass

        result['added_qty'] = int(result['added_qty']) if result['added_qty'].is_integer() else result['added_qty']
        result['cart_quantity'] = order.cart_quantity or 0
        request.session['website_sale_cart_quantity'] = order.cart_quantity
        return result

    def _import_cart_lines(self, rows):
        result = self._empty_import_result()
        order = request.website.sale_get_order(force_create=True, update_pricelist=False)
        Product = request.env['product.product'].sudo()

        for row_index, code, qty_value in rows:
            code = str(code or '').strip()
            if not code and (qty_value in (None, '')):
                continue

            try:
                qty = float(qty_value or 0)
            except Exception:
                qty = 0

            if not code:
                result['failed_count'] += 1
                result['failures'].append({'row': row_index, 'code': '', 'reason': 'Thiếu mã SP'})
                continue
            if qty <= 0:
                result['failed_count'] += 1
                result['failures'].append({'row': row_index, 'code': code, 'reason': 'Số lượng phải > 0'})
                continue

            product = Product.search([('default_code', '=', code)], limit=1)
            if not product:
                product = Product.search([('barcode', '=', code)], limit=1)
            if not product or not product.sale_ok:
                result['failed_count'] += 1
                result['failures'].append({'row': row_index, 'code': code, 'reason': 'Không tìm thấy sản phẩm bán được'})
                continue

            try:
                values = order._cart_update(product_id=product.id, add_qty=qty) or {}
                warning = values.get('warning') or ''
                if warning:
                    result['failed_count'] += 1
                    result['failures'].append({'row': row_index, 'code': code, 'reason': warning})
                    continue
                result['success_count'] += 1
                result['added_qty'] += qty
            except Exception as error:
                _logger.exception('Import cart row failed: row=%s code=%s', row_index, code)
                result['failed_count'] += 1
                result['failures'].append({'row': row_index, 'code': code, 'reason': str(error)})

        return self._finish_import_result(order, result)

    @http.route('/shop/import_cart_template.xlsx', type='http', auth='public', website=True, sitemap=False)
    def download_import_cart_template(self, **_kwargs):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet('Import Cart')
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#2563EB',
            'font_color': '#FFFFFF',
            'border': 1,
        })
        text_format = workbook.add_format({'border': 1})
        qty_format = workbook.add_format({'border': 1, 'num_format': '0'})

        sheet.set_column('A:A', 22)
        sheet.set_column('B:B', 14)
        sheet.write(0, 0, 'MaSP', header_format)
        sheet.write(0, 1, 'SoLuong', header_format)
        sheet.write(1, 0, 'SP001', text_format)
        sheet.write(1, 1, 1, qty_format)
        sheet.data_validation('B2:B1000', {'validate': 'integer', 'criteria': '>=', 'value': 1})
        workbook.close()
        output.seek(0)

        headers = [
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', 'attachment; filename="mau_import_gio_hang.xlsx"'),
        ]
        return request.make_response(output.read(), headers=headers)

    @http.route('/shop/import_cart_excel', type='json', auth='public', website=True, csrf=False)
    def import_cart_excel(self, file=None, filename=None, **_kwargs):
        result = self._empty_import_result()
        if not file:
            result['failed_count'] = 1
            result['failures'].append({'row': 0, 'code': '', 'reason': 'Chưa chọn file'})
            return result

        try:
            raw = base64.b64decode(file.split(',', 1)[-1])
            workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = workbook.active
        except Exception:
            result['failed_count'] = 1
            result['failures'].append({'row': 0, 'code': filename or '', 'reason': 'File Excel không hợp lệ'})
            return result

        rows = (
            (row_index, row[0] if row and len(row) > 0 else '', row[1] if row and len(row) > 1 else None)
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2)
        )
        return self._import_cart_lines(rows)

    @http.route('/shop/import_cart_text', type='json', auth='public', website=True, csrf=False)
    def import_cart_text(self, text=None, **_kwargs):
        result = self._empty_import_result()
        text = (text or '').strip()
        if not text:
            result['failed_count'] = 1
            result['failures'].append({'row': 0, 'code': '', 'reason': 'Chưa nhập danh sách sản phẩm'})
            return result

        rows = []
        for index, chunk in enumerate(text.split(';'), start=1):
            chunk = chunk.strip()
            if not chunk:
                continue
            parts = [part.strip() for part in chunk.split(',')]
            if len(parts) < 2:
                rows.append((index, chunk, 0))
                continue
            rows.append((index, parts[0], parts[1]))
        if not rows:
            result['failed_count'] = 1
            result['failures'].append({'row': 0, 'code': '', 'reason': 'Không có cặp mã SP, số lượng hợp lệ'})
            return result
        return self._import_cart_lines(rows)
